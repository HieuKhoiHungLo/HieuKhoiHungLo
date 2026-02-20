import openpyxl

wb = openpyxl.load_workbook("d:/xampp/htdocs/TS/database/geodata_new.xlsx", data_only=True)
print("Sheet names found:")
for name in wb.sheetnames:
    print(f"'{name}'")

# Based on previous output, let's try to match by index or substring
# Usually sheet 2 is the one we want if it contains "sau sáp nhập"
sheet2_name = [n for n in wb.sheetnames if "2." in n][0]
sheet = wb[sheet2_name]
print(f"Targeting sheet: '{sheet2_name}'")

for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=50, values_only=True)):
    print(f"Row {i+1}: {row}")

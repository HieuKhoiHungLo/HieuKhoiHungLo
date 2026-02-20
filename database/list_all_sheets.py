import openpyxl

wb = openpyxl.load_workbook("d:/xampp/htdocs/TS/database/geodata_new.xlsx", data_only=True)
print("TOTAL SHEETS:", len(wb.sheetnames))
for i, name in enumerate(wb.sheetnames):
    print(f"Sheet {i}: '{name}' (len {len(name)})")

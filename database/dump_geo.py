import openpyxl

wb = openpyxl.load_workbook("d:/xampp/htdocs/TS/database/geodata_new.xlsx", data_only=True)
print(f"Sheets: {wb.sheetnames}")
sheet = wb.active # Let's see which one is active
print(f"Active Sheet: {sheet.title}")

for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=50, values_only=True)):
    print(f"Row {i+1}: {row}")

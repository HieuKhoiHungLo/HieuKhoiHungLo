import openpyxl

wb = openpyxl.load_workbook("d:/xampp/htdocs/TS/database/geodata_new.xlsx", data_only=True)
sheet = wb.active

# Look for District and Province patterns
# Some files have Province and District as headers above their respective blocks of wards
for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=200, values_only=True)):
    data = [str(c) if c is not None else "None" for c in row]
    
    # Check if a row has very few non-None values, it might be a header for a section
    non_none = [v for v in data if v != "None"]
    if 1 <= len(non_none) <= 5:
        print(f"Row {i+1} (Possible Header): {data}")
    
    # Check for keywords
    row_str = " ".join(data)
    if any(k in row_str for k in ["Quận", "Huyện", "Thị xã", "Tỉnh"]):
        print(f"Row {i+1} (Keyword Match): {data}")

# Let's also check column headers specifically
print("\n--- Rows 1-10 raw ---")
for i in range(1, 11):
    print(f"Row {i}: {[sheet.cell(i, j).value for j in range(1, 15)]}")
